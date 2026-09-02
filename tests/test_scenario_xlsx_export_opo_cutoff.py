"""Выгрузка «Бухгалтерия»: колонка ОПЭ пропадает с квартала отсечки."""

from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.models import (
    AppSetting, BacklogItem, Employee, EmployeeTeam, PlanningScenario,
    Role, ScenarioAllocation,
)
from app.services import opo_policy
from app.services.scenario_xlsx_export import ScenarioXlsxExporter


@pytest.fixture
def scenario_with_opo(db_session):
    """Сценарий Q4 2026 с одной включённой задачей: 20 АН / 30 ПР / 10 ТС / 10 ОПЭ."""
    db_session.add_all([
        Role(code="dev", label="Разработчик", color="#1890FF",
             is_active=True, counts_in_planning=True),
        Role(code="analyst", label="Аналитик", color="#722ED1",
             is_active=True, counts_in_planning=True),
    ])
    db_session.flush()
    emp = Employee(jira_account_id="o1", display_name="Olga", role="analyst", is_active=True)
    db_session.add(emp)
    db_session.flush()
    db_session.add(EmployeeTeam(employee_id=emp.id, team="Omega", is_primary=True))

    item = BacklogItem(
        title="ОПЭ feature", priority=1,
        estimate_hours=70, estimate_analyst_hours=20, estimate_dev_hours=30,
        estimate_qa_hours=10, estimate_opo_hours=10, opo_analyst_ratio=0.5,
    )
    db_session.add(item)
    db_session.flush()

    scenario = PlanningScenario(
        name="Q4 2026 Omega", year=2026, quarter="Q4", team="Omega", status="draft",
    )
    db_session.add(scenario)
    db_session.flush()
    db_session.add(ScenarioAllocation(
        scenario_id=scenario.id, backlog_item_id=item.id,
        included_flag=True, planned_hours=70.0,
    ))
    db_session.flush()
    return scenario.id


def _included_sheet(db_session, scenario_id):
    data = ScenarioXlsxExporter(db_session, scenario_id).build()
    return load_workbook(BytesIO(data))["Включено"]


def _headers(ws) -> list:
    return [ws.cell(row=2, column=c).value for c in range(1, 12)]


def test_opo_column_present_without_cutoff(db_session, scenario_with_opo):
    ws = _included_sheet(db_session, scenario_with_opo)
    assert "ОПЭ, ч" in _headers(ws)


def test_total_does_not_double_count_opo(db_session, scenario_with_opo):
    ws = _included_sheet(db_session, scenario_with_opo)
    headers = _headers(ws)
    total_col = headers.index("Итого, ч") + 1
    assert ws.cell(row=3, column=total_col).value == 70.0


def test_opo_column_dropped_from_cutoff(db_session, scenario_with_opo):
    db_session.add(AppSetting(key=opo_policy.SETTING_KEY, value="2026Q4"))
    db_session.flush()

    ws = _included_sheet(db_session, scenario_with_opo)
    headers = _headers(ws)

    assert "ОПЭ, ч" not in headers
    assert ws.cell(row=3, column=5).value == 25.0  # анализ 20 + половина ОПЭ
    assert ws.cell(row=3, column=6).value == 35.0  # разработка 30 + половина ОПЭ
    assert ws.cell(row=3, column=7).value == 10.0  # тестирование
    assert ws.cell(row=3, column=8).value == 70.0  # итого
