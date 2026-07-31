"""Выгрузка отчёта KPI в Excel — числа округлены так же, как их показывает интерфейс."""
from io import BytesIO

from openpyxl import load_workbook

from app.services.kpi.xlsx_export import export_report_xlsx


def _report(rows: list[dict]) -> dict:
    return {"year": 2026, "month": 7, "rows": rows}


def test_metric_value_rounded_like_the_screen():
    """Находка 5: 59.99999999999999 в книге должно стать 60, как на экране (``Math.round``)."""
    report = _report([{
        "team": "Платежи", "employee_name": "Иванов И.",
        "metrics": [{"name": "Своевременность трудозатрат", "value": 59.99999999999999, "has_data": True}],
        "total": None,
    }])
    wb = load_workbook(BytesIO(export_report_xlsx(report)))
    ws = wb.active
    row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    assert row[2] == 60


def test_total_rounded_like_the_screen():
    report = _report([{
        "team": "Платежи", "employee_name": "Иванов И.",
        "metrics": [], "total": 86.54999999999998,
    }])
    wb = load_workbook(BytesIO(export_report_xlsx(report)))
    ws = wb.active
    row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    assert row[-1] == 87


def test_no_data_metric_still_written_as_text_not_number():
    report = _report([{
        "team": "Платежи", "employee_name": "Иванов И.",
        "metrics": [{"name": "Качество выпуска", "value": None, "has_data": False}],
        "total": None,
    }])
    wb = load_workbook(BytesIO(export_report_xlsx(report)))
    ws = wb.active
    row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    assert row[2] == "нет данных"
