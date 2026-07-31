"""Тесты API раздела KPI: отчёт, сводка по командам, расшифровка, тренд, утверждение месяца."""

import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.main import app
from app.database import Base, get_db
from app.models import AppSetting, Employee, EmployeeTeam, Issue, Project
from app.services.kpi.seed import seed_defaults


@pytest.fixture
def db_session():
    """StaticPool in-memory DB so the TestClient thread shares the same connection."""
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(bind=engine)
    TestingSession = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    session = TestingSession()
    try:
        yield session
    finally:
        session.close()
        engine.dispose()


@pytest.fixture
def client(db_session):
    def _get_db():
        yield db_session
    app.dependency_overrides[get_db] = _get_db
    try:
        yield TestClient(app)
    finally:
        app.dependency_overrides.clear()


@pytest.fixture
def team_with_analyst(db_session):
    """Команда «Платежи» с одним аналитиком и профилем оценки по умолчанию."""
    seed_defaults(db_session)
    db_session.commit()

    project = Project(jira_project_id="1", key="OS", name="1С")
    db_session.add(project)
    emp = Employee(jira_account_id="acc-1", display_name="Иванов И.", team="Платежи", role="analyst")
    db_session.add(emp)
    db_session.commit()
    db_session.add(EmployeeTeam(employee_id=emp.id, team="Платежи", is_primary=True))
    db_session.commit()
    return {"project": project, "employee": emp}


class TestReport:
    @pytest.mark.no_auth_bypass
    def test_report_requires_auth(self, client):
        resp = client.get("/api/v1/kpi/report?year=2026&month=7")
        assert resp.status_code in (401, 403)

    def test_report_returns_rows_and_summary(self, client, team_with_analyst):
        resp = client.get("/api/v1/kpi/report?year=2026&month=7&teams=Платежи")
        assert resp.status_code == 200, resp.text
        body = resp.json()
        assert "rows" in body and "summary" in body
        assert [r["employee_name"] for r in body["rows"]] == ["Иванов И."]
        assert "avg_total" in body["summary"]

    def test_report_defaults_to_all_teams_when_omitted(self, client, team_with_analyst):
        resp = client.get("/api/v1/kpi/report?year=2026&month=7")
        assert resp.status_code == 200
        assert [r["employee_name"] for r in resp.json()["rows"]] == ["Иванов И."]


class TestConditionErrorNotA500:
    """BLOCKER 2: опечатка в уже сохранённых условиях — понятный 422, а не 500."""

    def test_report_returns_422_when_metric_has_bad_condition(
        self, client, db_session, team_with_analyst
    ):
        import json

        from app.models.kpi import KpiMetric, KpiProfile, KpiProfileMetric

        profile = db_session.query(KpiProfile).filter_by(code="analyst").one()
        bad_metric = KpiMetric(
            code="broken", name="Сломанная метрика", calc_kind="ratio",
            numerator_json=json.dumps({
                "unit": "issues", "person_field": "author", "period_window": "closed_in",
                "conditions": [{"attr": "environmentt", "op": "eq", "value": "PROD"}],
            }),
            denominator_json=json.dumps({"conditions": []}),
        )
        db_session.add(bad_metric)
        db_session.commit()
        db_session.add(KpiProfileMetric(profile_id=profile.id, metric_id=bad_metric.id, weight=0.0))
        db_session.commit()

        resp = client.get("/api/v1/kpi/report?year=2026&month=7&teams=Платежи")
        assert resp.status_code == 422, resp.text
        assert "атрибут" in resp.json()["detail"]


class TestTeamsSummary:
    def test_teams_summary_lists_team_with_delta(self, client, team_with_analyst):
        resp = client.get("/api/v1/kpi/teams-summary?year=2026&month=7")
        assert resp.status_code == 200, resp.text
        rows = resp.json()["rows"]
        assert any(r["team"] == "Платежи" for r in rows)
        row = next(r for r in rows if r["team"] == "Платежи")
        assert "delta" in row
        assert "member_count" in row and "metrics" in row

    def test_teams_summary_scoped_to_filter(self, client, db_session, team_with_analyst):
        """ВАЖНО 11: сводка считает только команды из фильтра, не все команды сервиса."""
        from app.models.employee import Employee
        from app.models.employee_team import EmployeeTeam

        outsider = Employee(jira_account_id="acc-9", display_name="Сидоров С.",
                            team="Другая команда", role="analyst")
        db_session.add(outsider)
        db_session.commit()
        db_session.add(EmployeeTeam(employee_id=outsider.id, team="Другая команда", is_primary=True))
        db_session.commit()

        resp = client.get("/api/v1/kpi/teams-summary?year=2026&month=7&teams=Платежи")
        assert resp.status_code == 200, resp.text
        teams_in_response = {r["team"] for r in resp.json()["rows"]}
        assert teams_in_response == {"Платежи"}


class TestBreakdown:
    def test_breakdown_unknown_metric_404(self, client, team_with_analyst):
        resp = client.get(
            "/api/v1/kpi/breakdown"
            "?account_id=acc-1&metric_code=does-not-exist&year=2026&month=7"
        )
        assert resp.status_code == 404

    def test_breakdown_returns_task_lists_with_jira_links(self, client, db_session, team_with_analyst):
        project = team_with_analyst["project"]
        from datetime import datetime

        released = Issue(
            jira_issue_id="r1", key="OS-100", summary="Задача", issue_type="Задача",
            status="ГОТОВО", status_category="done", resolution="Готово",
            resolved_at=datetime(2026, 7, 10), project_id=project.id,
            reporter_account_id="acc-1", team="Платежи",
        )
        db_session.add(released)
        db_session.add(AppSetting(key="jira_base_url", value="https://itgri.atlassian.net"))
        db_session.commit()

        resp = client.get(
            "/api/v1/kpi/breakdown"
            "?account_id=acc-1&metric_code=quality&year=2026&month=7&teams=Платежи"
        )
        assert resp.status_code == 200, resp.text
        body = resp.json()
        assert body["metric_code"] == "quality"
        assert any(t["key"] == "OS-100" for t in body["denominator"])
        item = next(t for t in body["denominator"] if t["key"] == "OS-100")
        assert item["url"] == "https://itgri.atlassian.net/browse/OS-100"

    def test_breakdown_worklog_items_have_distinct_ids_for_same_task(
        self, client, db_session, team_with_analyst,
    ):
        """Находка 3: у одной задачи бывает несколько списаний трудозатрат —
        расшифровка «Своевременности трудозатрат» должна ключевать записи по
        идентификатору списания, а не по задаче, иначе список даёт
        повторяющиеся ключи."""
        from datetime import datetime

        from app.models.worklog import Worklog

        project = team_with_analyst["project"]
        emp = team_with_analyst["employee"]

        wl_issue = Issue(
            jira_issue_id="wl-host", key="OS-1500", summary="Задача", issue_type="Задача",
            status="ГОТОВО", status_category="done", project_id=project.id, team="Платежи",
        )
        db_session.add(wl_issue)
        db_session.commit()
        db_session.add_all([
            Worklog(
                jira_worklog_id="wl-a", issue_id=wl_issue.id, employee_id=emp.id,
                started_at=datetime(2026, 7, 7, 10, 0),
                jira_created_at=datetime(2026, 7, 7, 18, 0),
                hours=1.0, time_spent_seconds=3600,
            ),
            Worklog(
                jira_worklog_id="wl-b", issue_id=wl_issue.id, employee_id=emp.id,
                started_at=datetime(2026, 7, 8, 10, 0),
                jira_created_at=datetime(2026, 7, 8, 18, 0),
                hours=2.0, time_spent_seconds=7200,
            ),
        ])
        db_session.commit()

        resp = client.get(
            "/api/v1/kpi/breakdown"
            "?account_id=acc-1&metric_code=worklog_timeliness&year=2026&month=7&teams=Платежи"
        )
        assert resp.status_code == 200, resp.text
        body = resp.json()
        assert len(body["denominator"]) == 2
        ids = [item["id"] for item in body["denominator"]]
        assert len(set(ids)) == 2
        assert all(item["key"] == "OS-1500" for item in body["denominator"])


class TestTrend:
    def test_trend_returns_requested_number_of_points(self, client, team_with_analyst):
        resp = client.get(
            "/api/v1/kpi/trend?account_id=acc-1&year=2026&month=7&months=3&teams=Платежи"
        )
        assert resp.status_code == 200, resp.text
        points = resp.json()["points"]
        assert len(points) == 3
        assert [p["month"] for p in points] == [5, 6, 7]

    def test_trend_unknown_employee_404(self, client, team_with_analyst):
        resp = client.get("/api/v1/kpi/trend?account_id=missing&year=2026&month=7")
        assert resp.status_code == 404


class TestApproval:
    def test_approve_and_read_back(self, client, team_with_analyst):
        resp = client.post(
            "/api/v1/kpi/approve", json={"team": "Платежи", "year": 2026, "month": 7}
        )
        assert resp.status_code == 200, resp.text
        assert resp.json()["approved_by"]

        got = client.get("/api/v1/kpi/approval?team=Платежи&year=2026&month=7")
        assert got.status_code == 200
        assert got.json()["approved"] is True
        assert got.json()["approved_by"]

    def test_approval_not_yet_approved(self, client, team_with_analyst):
        resp = client.get("/api/v1/kpi/approval?team=Платежи&year=2026&month=7")
        assert resp.status_code == 200
        assert resp.json()["approved"] is False

    def test_reapprove_overwrites_snapshot_without_error(self, client, team_with_analyst):
        first = client.post(
            "/api/v1/kpi/approve", json={"team": "Платежи", "year": 2026, "month": 7}
        )
        assert first.status_code == 200
        second = client.post(
            "/api/v1/kpi/approve", json={"team": "Платежи", "year": 2026, "month": 7}
        )
        assert second.status_code == 200


class TestExport:
    def test_export_xlsx_returns_workbook(self, client, team_with_analyst):
        resp = client.get("/api/v1/kpi/export.xlsx?year=2026&month=7&teams=Платежи")
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert len(resp.content) > 0

    def test_export_rows_match_report_rows(self, client, team_with_analyst):
        """ВАЖНО 8: выгрузка открывается как книга, её строки совпадают со строками отчёта."""
        from io import BytesIO
        from openpyxl import load_workbook

        report = client.get("/api/v1/kpi/report?year=2026&month=7&teams=Платежи").json()
        resp = client.get("/api/v1/kpi/export.xlsx?year=2026&month=7&teams=Платежи")
        assert resp.status_code == 200

        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        names_in_sheet = [row[1] for row in ws.iter_rows(min_row=2, values_only=True)]
        assert names_in_sheet == [r["employee_name"] for r in report["rows"]]


class TestApprovalFreeze:
    """BLOCKER 1: утверждённый месяц не меняется после правки весов профиля."""

    def test_report_frozen_after_weight_change(self, client, db_session, team_with_analyst):
        from datetime import datetime

        from app.models import AppSetting
        from app.models.kpi import KpiProfile
        from app.services.kpi.kpi_service import build_report

        project = team_with_analyst["project"]
        # empty_policy=zero делает вес метрики значимым для итога даже когда
        # данные есть только у одной метрики (redistribute свёл бы к тому же
        # числу независимо от веса — тест не отличил бы заморозку от совпадения).
        db_session.add(AppSetting(key="kpi_empty_policy", value="zero"))
        db_session.add(Issue(
            jira_issue_id="fr1", key="OS-900", summary="s", issue_type="Задача",
            status="ГОТОВО", status_category="done", resolution="Готово",
            resolved_at=datetime(2026, 7, 10), project_id=project.id,
            reporter_account_id="acc-1", team="Платежи",
        ))
        db_session.commit()

        before = client.get("/api/v1/kpi/report?year=2026&month=7&teams=Платежи").json()
        row_before = next(r for r in before["rows"] if r["account_id"] == "acc-1")
        assert row_before["total"] is not None

        approve = client.post("/api/v1/kpi/approve", json={"team": "Платежи", "year": 2026, "month": 7})
        assert approve.status_code == 200

        profile = db_session.query(KpiProfile).filter_by(code="analyst").one()
        quality_link = next(m for m in profile.metrics if m.metric.code == "quality")
        quality_link.weight = 0.9
        db_session.commit()

        # Живой пересчёт теперь дал бы другое число — иначе тест ничего не
        # проверяет (сравнивал бы совпадающие по случайности значения).
        live = build_report(db_session, ["Платежи"], 2026, 7)
        live_row = next(r for r in live["rows"] if r["account_id"] == "acc-1")
        assert live_row["total"] != row_before["total"]

        after = client.get("/api/v1/kpi/report?year=2026&month=7&teams=Платежи").json()
        row_after = next(r for r in after["rows"] if r["account_id"] == "acc-1")
        assert row_after["total"] == row_before["total"]
        assert after["approvals"]["Платежи"]["approved"] is True


class TestBreakdownConsistency:
    """BLOCKER 2: расшифровка использует тот же отбор, что и дробь в отчёте."""

    def test_breakdown_matches_report_fraction_for_mid_month_join(self, client, db_session):
        from datetime import date, datetime

        seed_defaults(db_session)
        db_session.commit()

        project = Project(jira_project_id="p-bd", key="OS", name="1С")
        db_session.add(project)
        emp = Employee(jira_account_id="acc-mid", display_name="Петров П.", team="Платежи", role="analyst")
        db_session.add(emp)
        db_session.commit()
        db_session.add(EmployeeTeam(
            employee_id=emp.id, team="Платежи", is_primary=True, joined_at=date(2026, 7, 20),
        ))
        # Задача, закрытая до вступления в команду — не должна попасть в знаменатель.
        db_session.add(Issue(
            jira_issue_id="bd1", key="OS-900", summary="до", issue_type="Задача",
            status="ГОТОВО", status_category="done", resolution="Готово",
            resolved_at=datetime(2026, 7, 5), project_id=project.id,
            reporter_account_id="acc-mid", team="Платежи",
        ))
        # Задача после вступления — попадает.
        db_session.add(Issue(
            jira_issue_id="bd2", key="OS-901", summary="после", issue_type="Задача",
            status="ГОТОВО", status_category="done", resolution="Готово",
            resolved_at=datetime(2026, 7, 25), project_id=project.id,
            reporter_account_id="acc-mid", team="Платежи",
        ))
        db_session.commit()

        report = client.get("/api/v1/kpi/report?year=2026&month=7&teams=Платежи").json()
        row = next(r for r in report["rows"] if r["account_id"] == "acc-mid")
        quality = next(m for m in row["metrics"] if m["code"] == "quality")
        assert quality["denominator"] == 1  # только задача от 25 июля

        breakdown = client.get(
            "/api/v1/kpi/breakdown"
            "?account_id=acc-mid&metric_code=quality&year=2026&month=7&teams=Платежи"
        ).json()
        assert breakdown["denominator_count"] == 1
        assert len(breakdown["denominator"]) == 1
        assert breakdown["denominator"][0]["key"] == "OS-901"


class TestDirections:
    def test_directions_lists_unique_values(self, client, db_session, team_with_analyst):
        project = team_with_analyst["project"]
        db_session.add(Issue(
            jira_issue_id="dir1", key="OS-800", summary="s", issue_type="Задача",
            status="ГОТОВО", status_category="done", project_id=project.id,
            team="Платежи", direction="Финансовые операции",
        ))
        db_session.commit()
        resp = client.get("/api/v1/kpi/directions")
        assert resp.status_code == 200
        assert "Финансовые операции" in resp.json()
