"""Выгрузки в xlsx: колонка «Группа» и разрез сценария по группам."""

from datetime import datetime
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.models import (
    Category,
    Employee,
    EmployeeTeam,
    Issue,
    Project,
    Role,
    Team,
    TeamSubgroup,
    Worklog,
)
from app.services.analytics_service import AnalyticsService
from app.services.export_service import ExportService
from app.services.subgroup_resolver import SubgroupResolver

TEAM = "Команда 1С (Бухгалтерия)"
DAY = datetime(2026, 8, 5, 10, 0)


@pytest.fixture
def data(db_session):
    db_session.add(Project(id="p1", jira_project_id="1", key="OS", name="OS"))
    db_session.add(Category(code="development", label="Разработка", color="#123456"))
    db_session.add(Role(code="dev", label="Программист", is_active=True, sort_order=1))
    team = Team(name=TEAM, has_subgroups=True)
    db_session.add(team)
    db_session.flush()
    calc = TeamSubgroup(team_id=team.id, name="Расчёты", sort_order=1)
    db_session.add(calc)
    emp = Employee(jira_account_id="acc-a", display_name="Алексеев", is_active=True, role="dev")
    db_session.add(emp)
    db_session.flush()
    db_session.add(EmployeeTeam(employee_id=emp.id, team=TEAM, is_primary=True, subgroup_id=calc.id))

    issue = Issue(
        jira_issue_id="OS-1", key="OS-1", summary="Задача", issue_type="Task",
        status="Open", project_id="p1", team=TEAM, category="development",
        assigned_subgroup_id=calc.id,
    )
    db_session.add(issue)
    db_session.flush()
    db_session.add(
        Worklog(jira_worklog_id="w1", issue_id=issue.id, employee_id=emp.id,
                started_at=DAY, time_spent_seconds=4 * 3600, hours=4.0)
    )
    db_session.commit()
    SubgroupResolver(db_session).recompute_effective()
    return {"calc": calc}


def _sheet(blob: bytes):
    return load_workbook(BytesIO(blob)).active


def test_analytics_export_has_group_column(db_session, data):
    report = AnalyticsService(db_session).get_hierarchical_report(
        year=2026, quarter=3, teams=[TEAM]
    )
    ws = _sheet(ExportService(db_session).export_analytics_report_xlsx(report, []))

    headers = [c.value for c in ws[1]]
    assert headers[1] == "Группа"

    rows = [[c.value for c in r] for r in ws.iter_rows(min_row=2)]
    assert any(r[1] == "Расчёты" for r in rows)


def test_group_column_empty_for_team_without_split(db_session, data):
    """У команды без деления состав колонок тот же, значение пустое."""
    from app.models import Team as TeamModel

    db_session.query(TeamModel).filter(TeamModel.name == TEAM).one().has_subgroups = False
    db_session.commit()
    SubgroupResolver(db_session).recompute_effective()

    report = AnalyticsService(db_session).get_hierarchical_report(
        year=2026, quarter=3, teams=[TEAM]
    )
    ws = _sheet(ExportService(db_session).export_analytics_report_xlsx(report, []))

    headers = [c.value for c in ws[1]]
    rows = [[c.value for c in r] for r in ws.iter_rows(min_row=2)]

    assert headers[1] == "Группа"
    assert all(not r[1] for r in rows)
