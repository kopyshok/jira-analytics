"""WorkTypeReportXlsx — 3-sheet export."""
import json
from io import BytesIO
from datetime import date, datetime

import pytest
from openpyxl import load_workbook

from app.models.work_type_report_snapshot import WorkTypeReportSnapshot
from app.models.mandatory_work_type import MandatoryWorkType
from app.services.work_type_report_xlsx import export_snapshot_to_xlsx


@pytest.fixture
def snap(db_session):
    wt = MandatoryWorkType(code="support_consult", label="Сопр", sort_order=1)
    db_session.add(wt); db_session.commit()
    data = {
        "headline": "AI summary text",
        "totals": {"hours": 540, "tasks": 78, "employees": 12, "themes_count": 1},
        "themes": [{
            "theme_id": "T1", "name": "Ошибки обмена", "color": "#00c9c8",
            "totals": {"hours": 173, "pct": 32, "tasks_count": 18, "employees_count": 5},
            "narrative": "Преобладали сбои в задачах PROJ-321...",
            "issues": [{
                "issue_id": "i1", "key": "PROJ-321", "summary": "Расхождения",
                "hours": 86, "contribution": "разбор расхождений",
                "employee_breakdown": [
                    {"name": "Иванов И.", "role": "analyst", "team": "Платформа", "hours": 60},
                    {"name": "Петров П.", "role": "dev", "team": "Платформа", "hours": 26},
                ],
            }],
        }],
        "candidates": [],
        "outliers": [],
        "recommendation": {"text": "Внедрить регламент", "expected_impact": "−80 ч/мес"},
        "manual_review_required": [],
        "is_fallback_narrative": False,
    }
    snap = WorkTypeReportSnapshot(
        work_type_id=wt.id, year=2026, quarter=2, month=4,
        start_date=date(2026, 4, 1), end_date=date(2026, 4, 30),
        team_set_hash="all", team_set_json=json.dumps([]),
        snapshot_data=json.dumps(data, ensure_ascii=False),
        dictionary_version=1, generated_at=datetime(2026, 5, 1),
    )
    db_session.add(snap); db_session.commit()
    return snap


def test_xlsx_three_sheets(snap):
    blob = export_snapshot_to_xlsx(snap)
    wb = load_workbook(BytesIO(blob))
    assert wb.sheetnames == ["Темы", "Задачи", "Текст"]


def test_xlsx_themes_sheet(snap):
    blob = export_snapshot_to_xlsx(snap)
    wb = load_workbook(BytesIO(blob))
    ws = wb["Темы"]
    headers = [c.value for c in ws[1]]
    assert headers == ["Тема", "Часов", "Доля, %", "Задач", "Сотрудников"]
    row2 = [c.value for c in ws[2]]
    assert row2[0] == "Ошибки обмена"
    assert row2[1] == 173
    assert row2[2] == 32


def test_xlsx_tasks_sheet_explodes_employee_breakdown(snap):
    blob = export_snapshot_to_xlsx(snap)
    wb = load_workbook(BytesIO(blob))
    ws = wb["Задачи"]
    headers = [c.value for c in ws[1]]
    assert headers == ["Тема", "Ключ", "Заголовок", "Сотрудник", "Роль", "Команда", "Часы", "Что делали"]
    # Two employees on one task → 2 rows
    rows = [tuple(c.value for c in row) for row in ws.iter_rows(min_row=2)]
    assert len(rows) == 2
    employees = [r[3] for r in rows]
    assert "Иванов И." in employees and "Петров П." in employees


def test_xlsx_text_sheet_includes_headline_and_recommendation(snap):
    blob = export_snapshot_to_xlsx(snap)
    wb = load_workbook(BytesIO(blob))
    ws = wb["Текст"]
    all_text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    assert "AI summary text" in all_text
    assert "Преобладали сбои" in all_text
    assert "Внедрить регламент" in all_text
    assert "−80 ч/мес" in all_text or "-80 ч/мес" in all_text  # encoding tolerance
