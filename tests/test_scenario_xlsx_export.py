"""Tests for ScenarioXlsxExporter — 7-sheet beautiful scenario export."""

from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.models import (
    BacklogItem, Employee, EmployeeTeam, MandatoryWorkType,
    PlanningScenario, ScenarioAllocation, Role,
)
from app.services.scenario_xlsx_export import ScenarioXlsxExporter


EXPECTED_SHEETS = [
    "Обложка",
    "Включено",
    "Не вошло",
    "По ролям",
    "По сотрудникам",
    "Правила",
    "Отсутствия",
]


@pytest.fixture
def minimal_scenario(db_session):
    db_session.add_all([
        Role(code="dev", label="Разработчик", color="#1890FF",
             is_active=True, counts_in_planning=True),
        MandatoryWorkType(code="org", label="Орг. вопросы",
                          is_active=True, subtracts_from_pool=True),
    ])
    db_session.flush()

    emp = Employee(
        jira_account_id="d1", display_name="Dave", role="dev", is_active=True,
    )
    db_session.add(emp)
    db_session.flush()
    db_session.add(EmployeeTeam(
        employee_id=emp.id, team="Alpha", is_primary=True,
    ))

    item_in = BacklogItem(
        title="Build feature", priority=1,
        estimate_hours=80, estimate_dev_hours=80,
    )
    item_out = BacklogItem(
        title="Skipped feature", priority=5,
        estimate_hours=200, estimate_dev_hours=200,
    )
    db_session.add_all([item_in, item_out])
    db_session.flush()

    scenario = PlanningScenario(
        name="Q2 2026 Alpha Base", year=2026, quarter="Q2",
        team="Alpha", status="draft",
    )
    db_session.add(scenario)
    db_session.flush()

    db_session.add_all([
        ScenarioAllocation(
            scenario_id=scenario.id, backlog_item_id=item_in.id,
            included_flag=True, planned_hours=80.0,
        ),
        ScenarioAllocation(
            scenario_id=scenario.id, backlog_item_id=item_out.id,
            included_flag=False, planned_hours=0.0,
        ),
    ])
    db_session.flush()

    class _R:
        pass
    r = _R()
    r.scenario_id = scenario.id
    return r


class TestScaffold:
    def test_workbook_has_seven_sheets_in_order(self, db_session, minimal_scenario):
        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        assert wb.sheetnames == EXPECTED_SHEETS

    def test_unknown_scenario_raises(self, db_session):
        with pytest.raises(ValueError, match="not found"):
            ScenarioXlsxExporter(db_session, "no-such-id").build()
