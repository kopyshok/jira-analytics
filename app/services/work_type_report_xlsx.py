"""XLSX export для тематического отчёта.

3 листа: «Темы» / «Задачи» / «Текст».
"""
import json
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from app.models.work_type_report_snapshot import WorkTypeReportSnapshot


_HEADER_FONT = Font(bold=True)


def _set_headers(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT


def export_snapshot_to_xlsx(snap: WorkTypeReportSnapshot) -> bytes:
    data = json.loads(snap.snapshot_data)
    wb = Workbook()

    # Sheet 1: Темы
    ws1 = wb.active
    ws1.title = "Темы"
    _set_headers(ws1, ["Тема", "Часов", "Доля, %", "Задач", "Сотрудников"])
    for t in data.get("themes", []) or []:
        tot = t.get("totals", {}) or {}
        ws1.append([
            t.get("name"),
            tot.get("hours", 0), tot.get("pct", 0),
            tot.get("tasks_count", 0), tot.get("employees_count", 0),
        ])

    # Sheet 2: Задачи
    ws2 = wb.create_sheet("Задачи")
    _set_headers(ws2, ["Тема", "Ключ", "Заголовок", "Сотрудник", "Роль", "Команда", "Часы", "Что делали"])
    for t in data.get("themes", []) or []:
        for i in t.get("issues", []) or []:
            breakdown = i.get("employee_breakdown") or [
                {"name": "", "role": "", "team": "", "hours": i.get("hours", 0)}
            ]
            for emp_row in breakdown:
                ws2.append([
                    t.get("name"), i.get("key"), i.get("summary"),
                    emp_row.get("name"), emp_row.get("role"), emp_row.get("team"),
                    emp_row.get("hours"), i.get("contribution"),
                ])

    # Sheet 3: Текст
    ws3 = wb.create_sheet("Текст")
    ws3.append(["AI-заголовок"])
    ws3.append([data.get("headline", "")])
    ws3.append([])
    ws3.append(["Нарративы по темам"])
    for t in data.get("themes", []) or []:
        ws3.append([t.get("name"), t.get("narrative", "")])
    ws3.append([])
    rec = data.get("recommendation") or {}
    ws3.append(["Рекомендация", rec.get("text", "")])
    ws3.append(["Ожидаемый эффект", rec.get("expected_impact", "")])

    # Wrap text in cells with long content
    for ws in (ws1, ws2, ws3):
        for col in ws.iter_cols(min_row=1):
            for cell in col:
                if cell.value is not None and isinstance(cell.value, str) and len(cell.value) > 60:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
